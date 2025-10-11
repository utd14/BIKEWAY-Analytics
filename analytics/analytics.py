import pandas as pd
import matplotlib.pyplot as plt
import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import get_engine

plt.style.use('seaborn-v0_8-darkgrid')

CHARTS_DIR = 'charts'
os.makedirs(CHARTS_DIR, exist_ok=True)

engine = get_engine()

print("BIKEWAY Analytics - Visualisations")
print()

# 1. PIE CHART
query1 = """
SELECT order_status,
       COUNT(*) as order_count,
       ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER (), 2) as percentage
FROM orders
GROUP BY order_status
ORDER BY order_count DESC;
"""

df1 = pd.read_sql(query1, engine)
print(f"\n1. PIE CHART - Order Status Distribution")
print(f"Rows retrieved: {len(df1)}")
print(df1)

plt.figure(figsize=(10, 8))
colors = ['#66b3ff', '#ff9999', '#99ff99', '#ffcc99']
plt.pie(df1['order_count'], labels=df1['order_status'], autopct='%1.1f%%', 
        colors=colors, startangle=90, explode=[0.05]*len(df1))
plt.title('Order Status Distribution',
          fontsize=14, fontweight='bold')
plt.legend(df1['order_status'], loc='best')
plt.savefig(f'{CHARTS_DIR}/1_pie_order_status.png', dpi=300, bbox_inches='tight')
plt.close()
print(f"- Shows how many orders are in each status\n(1: Pending, 2: Processing, 3: Rejected, 4: Completed)")

# 2. BAR CHART
query2 = """
SELECT s.store_name, 
       SUM(oi.quantity * oi.list_price * (1 - oi.discount)) as total_revenue
FROM orders o
JOIN order_items oi ON o.order_id = oi.order_id
JOIN stores s ON o.store_id = s.store_id
GROUP BY s.store_name
ORDER BY total_revenue DESC;
"""

df2 = pd.read_sql(query2, engine)
print(f"\n2. BAR CHART - Total Revenue by Store")
print(f"Rows retrieved: {len(df2)}")
print(df2)

plt.figure(figsize=(12, 6))
bars = plt.bar(df2['store_name'], df2['total_revenue'], color='steelblue', edgecolor='black', linewidth=1.2)
plt.xlabel('Store Name', fontsize=12, fontweight='bold')
plt.ylabel('Total Revenue ($)', fontsize=12, fontweight='bold')
plt.title('Total Revenue by Store', fontsize=16, fontweight='bold')
plt.xticks(rotation=45, ha='right')
plt.grid(axis='y', alpha=0.3)
for bar in bars:
    height = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2., height,
             f'${height:,.0f}', ha='center', va='bottom', fontsize=9)
plt.tight_layout()
plt.savefig(f'{CHARTS_DIR}/2_bar_revenue_by_store.png', dpi=300, bbox_inches='tight')
plt.close()
print(f"- Shows a store that generates the most revenue")

# 3. HORIZONTAL BAR CHART - Top 10 Best-Sellers
query3 = """
SELECT p.product_name, 
       b.brand_name,
       SUM(oi.quantity) as total_quantity_sold
FROM order_items oi
JOIN products p ON oi.product_id = p.product_id
JOIN brands b ON p.brand_id = b.brand_id
GROUP BY p.product_name, b.brand_name
ORDER BY total_quantity_sold DESC
LIMIT 10;
"""

df3 = pd.read_sql(query3, engine)
print(f"\n3. HORIZONTAL BAR CHART - Top 10 Best-Selling Products")
print(f"Rows retrieved: {len(df3)}")
print(df3)

plt.figure(figsize=(12, 8))
df3_sorted = df3.sort_values('total_quantity_sold', ascending=True)
bars = plt.barh(df3_sorted['product_name'], df3_sorted['total_quantity_sold'], 
                color='coral', edgecolor='black', linewidth=1)
plt.xlabel('Total Quantity Sold', fontsize=12, fontweight='bold')
plt.ylabel('Product Name', fontsize=12, fontweight='bold')
plt.title('Top 10 Best-Selling Products', fontsize=16, fontweight='bold')
plt.grid(axis='x', alpha=0.3)
for i, bar in enumerate(bars):
    width = bar.get_width()
    plt.text(width, bar.get_y() + bar.get_height()/2., 
             f'{int(width)} units', ha='left', va='center', fontsize=9)
plt.tight_layout()
plt.savefig(f'{CHARTS_DIR}/3_horizontal_bar_top_products.png', dpi=300, bbox_inches='tight')
plt.close()
print(f"- Identifies most popular products by sales")

# 4. LINE CHART 
query4 = """
SELECT DATE_TRUNC('month', order_date) as month,
       COUNT(o.order_id) as total_orders,
       SUM(oi.quantity * oi.list_price * (1 - oi.discount)) as monthly_revenue
FROM orders o
JOIN order_items oi ON o.order_id = oi.order_id
GROUP BY DATE_TRUNC('month', order_date)
ORDER BY month;
"""

df4 = pd.read_sql(query4, engine)
print(f"\n4. LINE CHART - Monthly Sales Trend")
print(f"Rows retrieved: {len(df4)}")
print(df4.head(10))

plt.figure(figsize=(14, 6))
plt.plot(df4['month'], df4['monthly_revenue'], marker='o', linewidth=2, 
         markersize=6, color='green', label='Monthly Revenue')
plt.xlabel('Month', fontsize=12, fontweight='bold')
plt.ylabel('Revenue ($)', fontsize=12, fontweight='bold')
plt.title('Monthly Sales Trend', fontsize=16, fontweight='bold')
plt.xticks(rotation=45, ha='right')
plt.grid(True, alpha=0.3)
plt.legend(loc='best')
plt.tight_layout()
plt.savefig(f'{CHARTS_DIR}/4_line_monthly_sales.png', dpi=300, bbox_inches='tight')
plt.close()
print(f"- Shows sales performance over time")

# 5. HISTOGRAM
query5 = """
SELECT b.brand_name,
       AVG(oi.list_price) as avg_price
FROM order_items oi
JOIN products p ON oi.product_id = p.product_id
JOIN brands b ON p.brand_id = b.brand_id
GROUP BY b.brand_name;
"""

df5 = pd.read_sql(query5, engine)
print(f"\n5. Histogram - Brand Popularity Analysis")
print(f"Rows retrieved: {len(df5)}")
print(df5)

plt.figure(figsize=(12, 6))
plt.hist(df5['avg_price'], bins=15, color='purple', edgecolor='black', alpha=0.7)
plt.xlabel('Average Price ($)', fontsize=12, fontweight='bold')
plt.ylabel('Number of Brands', fontsize=12, fontweight='bold')
plt.gca().yaxis.set_major_locator(plt.MaxNLocator(integer=True))
plt.title('Brand Popularity Analysis', fontsize=16, fontweight='bold')
plt.grid(axis='y', alpha=0.3)
plt.tight_layout()
plt.savefig(f'{CHARTS_DIR}/5_histogram_brand_prices.png', dpi=300, bbox_inches='tight')
plt.close()
print(f"- Compares sales performance across brands (price distribution)")

# 6. SCATTER PLOT
query6 = """
SELECT s.store_name,
       COUNT(o.order_id) as total_orders,
       AVG(oi.quantity * oi.list_price * (1 - oi.discount)) as avg_order_value
FROM orders o
JOIN order_items oi ON o.order_id = oi.order_id
JOIN stores s ON o.store_id = s.store_id
GROUP BY s.store_name
ORDER BY avg_order_value DESC;
"""

df6 = pd.read_sql(query6, engine)
print(f"\n6. Scatter Plot - Average Order Value by Store")
print(f"Rows retrieved: {len(df6)}")
print(df6)

plt.figure(figsize=(12, 8))
plt.scatter(df6['total_orders'], df6['avg_order_value'], 
            s=200, alpha=0.6, color='orange', edgecolors='black', linewidth=1.5)
for i, store in enumerate(df6['store_name']):
    plt.annotate(store, (df6['total_orders'].iloc[i], df6['avg_order_value'].iloc[i]),
                 fontsize=10, ha='center')
plt.xlabel('Total Number of Orders', fontsize=12, fontweight='bold')
plt.ylabel('Average Order Value ($)', fontsize=12, fontweight='bold')
plt.title('Average Order Value by Store', fontsize=16, fontweight='bold')
plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f'{CHARTS_DIR}/6_scatter_store_performance.png', dpi=300, bbox_inches='tight')
plt.close()
print(f"- Calculates average revenue per item")

print()
print("Success")

# PLOTLY - Interactive Time Slider
import plotly.express as px

print()
print("Plotly Interactive Time Slider")
print()

# Daily revenue by month for each store
query_plotly = """
SELECT DATE_TRUNC('month', o.order_date)::date as month,
       s.store_name,
       COUNT(o.order_id) as monthly_orders,
       SUM(oi.quantity * oi.list_price * (1 - oi.discount)) as monthly_revenue
FROM orders o
JOIN order_items oi ON o.order_id = oi.order_id
JOIN stores s ON o.store_id = s.store_id
WHERE o.order_date IS NOT NULL
GROUP BY DATE_TRUNC('month', o.order_date)::date, s.store_name
ORDER BY month, store_name;
"""

df_plotly = pd.read_sql(query_plotly, engine)
print(f"\nRows retrieved: {len(df_plotly)}")
print(df_plotly.head(10))

fig = px.scatter(df_plotly, 
                 x="monthly_orders", 
                 y="monthly_revenue",
                 animation_frame="month",
                 animation_group="store_name",
                 size="monthly_revenue",
                 color="store_name",

                 hover_name="store_name",
                 size_max=60,
                 range_x=[0, df_plotly['monthly_orders'].max() + 20],
                 range_y=[0, df_plotly['monthly_revenue'].max() * 1.1],
                 title="Monthly Sales Performance by Store Over Time (All Years)",
                 labels={"monthly_orders": "Number of Orders", 
                         "monthly_revenue": "Monthly Revenue ($)",
                         "store_name": "Store"})

fig.update_layout(
    width=1200,
    height=700,
    font=dict(size=14),
    title_font=dict(size=18, family="Arial Black")
)

fig.show()

# print("Shows sales performance over time (montly) for each store")

print()

# Excelt expot
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

EXPORTS_DIR = 'exports'
os.makedirs(EXPORTS_DIR, exist_ok=True)
print("Excelt export + formatting")
def export_to_excel(dataframes_dict, filename):
    """
    Export multiple dataframes to Excel with formatting:
    - Frozen headers
    - Gradient fill for numeric columns
    - Auto filters
    - Conditional formatting (min/max highlighting)
    """
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    # I had some timezone issue related to Excel's incompatibility
    dataframes_fixed = {}
    for sheet_name, df in dataframes_dict.items():
        df_copy = df.copy()
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                df_copy[col] = pd.to_datetime(df_copy[col]).dt.tz_localize(None)
        dataframes_fixed[sheet_name] = df_copy
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_fixed.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    wb = load_workbook(filepath)
    
    total_rows = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        total_rows += (max_row - 1)  # minus the header
        
        # FRreeze panes (first row and first column)
        ws.freeze_panes = "B2"
        
        # Auto filter for all columns
        ws.auto_filter.ref = ws.dimensions
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 4. Find numeric columns and apply formatting
        numeric_cols = []
        for col_idx, cell in enumerate(ws[1], start=1):
            col_letter = ws.cell(1, col_idx).column_letter
            
            # Check if column contains numeric data (check first data row)
            if max_row > 1:
                first_value = ws.cell(2, col_idx).value
                if isinstance(first_value, (int, float)):
                    numeric_cols.append(col_letter)
        
        # 3. Gradient fill (ColorScaleRule) to numerics
        for col_letter in numeric_cols:
            col_range = f"{col_letter}2:{col_letter}{max_row}"
            
            # red (min) yellow (mid) green (max)
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )
            ws.conditional_formatting.add(col_range, rule)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    
    print(f"\nCreated file: {filename}, {len(dataframes_dict)} sheets, rows: {total_rows}")
    print(f"Location: {filepath}")
    
    return filepath
dataframes_to_export = {
    'Order Status': df1,
    'Revenue by Store': df2,
    'Top Products': df3,
    'Monthly Sales': df4,
    'Brand Prices': df5,
    'Store Performance': df6
}
export_to_excel(dataframes_to_export, 'bike_store_report.xlsx')
print("Exported to excel")